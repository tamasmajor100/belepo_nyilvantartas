import toga
from toga.style import Pack
from toga.style.pack import COLUMN, ROW
import openpyxl

class belepo(toga.App):

    def startup(self):

        # Adatok előkészítése
        self.data = []

        # Táblázat widget létrehozása
        self.table = toga.Table(
            data=self.data,
            headings=['Név', 'Nem', 'Életkor', 'Hely', 'Telephely'],
            style=Pack(flex=1)
        )

        #self.table2 = toga.Table(
        #    data=self.data,
        #    headings=['Név', 'Nem', 'Életkor', 'Hely', 'Telephely'],
        #    style=Pack(flex=1)
        #)

        # Importálás gomb

        nev_label = toga.Label("Név:")
        nev_label.style.update(flex=1, padding_left=27, padding_top=5, padding_bottom=3)
        self.nev_input = toga.TextInput(placeholder='Név')
        self.nev_input.style.update(flex=1, padding_left=30, padding_top=2, padding_right=30, padding_bottom=3)

        nem_label = toga.Label("Nem:")
        nem_label.style.update(flex=1, padding_left=27, padding_top=5, padding_bottom=3)
        self.nem_input = toga.TextInput(placeholder='Nem')
        self.nem_input.style.update(flex=1, padding_left=30, padding_top=2, padding_right=30, padding_bottom=3)

        eletkor_label = toga.Label("Életkor:")
        eletkor_label.style.update(flex=1, padding_left=27, padding_top=5, padding_bottom=3)
        self.eletkor_input = toga.TextInput(placeholder='Életkor')
        self.eletkor_input.style.update(flex=1, padding_left=30, padding_top=2, padding_right=30, padding_bottom=3)

        hely_label = toga.Label("Hely:")
        hely_label.style.update(flex=1, padding_left=27, padding_top=5, padding_bottom=3)
        self.hely_input = toga.TextInput(placeholder='Hely')
        self.hely_input.style.update(flex=1, padding_left=30, padding_top=2, padding_right=30, padding_bottom=3)

        telephely_label = toga.Label("Telephely:")
        telephely_label.style.update(flex=1, padding_left=27, padding_top=5, padding_bottom=3)
        self.telephely_input = toga.TextInput(placeholder='Telephely')
        self.telephely_input.style.update(flex=1, padding_left=30, padding_top=2, padding_right=30, padding_bottom=3)

        hozzaad_button = toga.Button("Dolgozó(k) felvétele",on_press=self.export_to_xlsx)
        hozzaad_button.style.update(flex=1, padding_left=30, padding_top=15, padding_right=30, padding_bottom=15)

        import_button = toga.Button('Adatok megtekintése', on_press=self.import_from_excel)
        import_button.style.update(flex=1, padding_top=15, padding_bottom=5)

        # Ablak elrendezése
        main_box = toga.Box(
            children=[#nev_label,
                      self.nev_input,
                      #nem_label,
                      self.nem_input,
                      #eletkor_label,
                      self.eletkor_input,
                      #hely_label,
                      self.hely_input,
                      #telephely_label,
                      self.telephely_input,
                      hozzaad_button,
                      #import_button,
                      self.table],
            style=Pack(direction='column', padding=10)
        )

        # Ablak létrehozása
        self.main_window = toga.MainWindow(title=self.formal_name, size=(500, 200))
        self.main_window.content = main_box

        # Beolvassa és megjeleníti az adatbázisban lévő adatokat
        self.import_from_excel(widget=main_box)

        self.main_window.show()

    def import_from_excel(self, widget):


        # Fájl kiválasztása
        #file_path = self.main_window.open_file_dialog('Import from Excel', ['.xlsx'])


        file_path = "C:/temp/adatbazis.xlsx"

        if file_path:
            # Adatok beolvasása Excel fájlból
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            imported_data = [list(row) for row in sheet.iter_rows(values_only=True)]

            # Ellenőrzés: legalább egy sor van-e
            if not imported_data:
                self.main_window.info_dialog('Hiba', 'Nem található adat az Excel fájlban.')
                return

            # Ellenőrzés: minden sor ugyanannyi oszlopot tartalmaz-e, mint az eredeti táblázat
            if len(imported_data[0]) != len(self.table.headings):
                self.main_window.info_dialog('Hiba',
                                             'Az Excel fájl oszlopainak száma nem egyezik a táblázattal.')
                return

            # Importált adatok hozzáadása a táblázathoz
            self.data.extend(imported_data)
            self.table.data = self.data

    def export_to_xlsx(self, widget):
        # Felhasználói adatok begyűjtése
        nev = self.nev_input.value
        nem = self.nem_input.value
        eletkor = self.eletkor_input.value
        hely = self.hely_input.value
        telephely = self.telephely_input.value

        # Üres mezők ellenőrzése
        if not nev or not nem or not eletkor or not hely or not telephely:
            self.main_window.info_dialog('HIBA!', 'Minden mezőt ki kell tölteni!')
            return

        # Felhasználói adatok hozzáadása a táblázathoz
        self.data.append([nev, nem, int(eletkor), hely, telephely])
        self.table.data = self.data

        # Input mezők ürítése
        self.nev_input.value = ''
        self.nem_input.value = ''
        self.eletkor_input.value = ''
        self.hely_input.value = ''
        self.telephely_input.value = ''


        # XLSX fájl létrehozása és adatok írása
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Oszlopfejlékek írása
        headings = self.data[0]
        for col_num, heading in enumerate(headings, 1):
            sheet.cell(row=1, column=col_num, value=heading)

        # Adatok írása
        for row_num, row_data in enumerate(self.data[1:], 2):
            for col_num, cell_value in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col_num, value=cell_value)

        # Fájl mentése
        #file_path = self.main_window.save_file_dialog('Export to XLSX', ['.xlsx'])
        file_path = "C:/temp/adatbazis.xlsx"

        if file_path:
            workbook.save(file_path)

        self.main_window.info_dialog('SIKERES MENTÉS!', 'Az adatok felvétele megtörtént!')




def main():
    return belepo()
